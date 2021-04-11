from Classes import Student
from openpyxl import load_workbook
from AppConfiguration import ClearTab, searchStudentsTab, colors, notebook, ROOT, tab4_SearchStudents
from ImportData import RemoveStrings
from tkinter import *
from tkinter import ttk
from tkinter.ttk import Treeview, Scrollbar
from tkinter import filedialog, messagebox
import os


fontSettings = ("Times New Roman", 16, 'bold')
txtWidth = 10
rowsNumber = 4
loadedWorkBook = None
infoColum =5

def DisplaySerachScreen():
    # label Title
    label1 = Label(searchStudentsTab, text="-- البحث عن أحد الطلاب --", background=colors.bg, foreground=colors.bg2,
                   padx=80, pady=15, relief=RIDGE,  font=("Times New Roman", 30, 'bold'))
    label1.grid(row=0, column=0, pady=15, columnspan=10, padx=50)

     # label Title
    label1 = Label(searchStudentsTab, text=" قم باختيار الملف الذي تريد البحث فيه بشرط أن يكون أحد الملفات المنتجة بواسطة هذا البرنامج",
                   background=colors.bg, pady=15, font=fontSettings, wraplength=400)
    label1.grid(row=1, column=0, pady=infoColum, columnspan=10, padx=150)

    btnAddExcelFile = Button(searchStudentsTab, text=" إضافة ملف الإكسل المراد البحث فيه ",
                             font=fontSettings, bg=colors.bg2, foreground="white", border=2,
                              cursor="hand2",command=lambda: AddExcelFileForSearch())
    btnAddExcelFile.grid(row=2, column=infoColum, pady=5, columnspan=3)



    label = Label(searchStudentsTab, text=':اسم الطالب', background=colors.bg,
                  foreground=colors.txtColor, font=fontSettings)
    label.grid(row=rowsNumber, column=8, sticky='E', padx=10)

    studentTxtBox = Entry(
        searchStudentsTab, font=fontSettings, width=txtWidth*4)
    studentTxtBox.configure(justify='right')
    studentTxtBox.insert(0, 'Fryd')
    studentTxtBox.grid(row=rowsNumber, column=4,
                       columnspan=3, padx=5, pady=10)

    btnSearchStudent = Button(
        searchStudentsTab, text="بحث", font=fontSettings, bg=colors.bg2,
        foreground="white", border=2, cursor="hand2",
        command=lambda: BtnSearchStudent(studentTxtBox.get()))
    btnSearchStudent.grid(row=rowsNumber, column=2, pady=10, padx=10)



def AddExcelFileForSearch():
    global loadedWorkBook
    importFilePath = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")])
    onlyFileName = os.path.split(importFilePath)
    label = Label(searchStudentsTab, text='File: ' + onlyFileName[1], background=colors.bg,
                  foreground=colors.txtColor, font=fontSettings)
    label.grid(row=3, column=infoColum, pady=5, columnspan=3)

    loadedWorkBook = load_workbook(importFilePath)



def BtnSearchStudent(studentName):
    if(loadedWorkBook == None):
        messagebox.showerror('error','برجاء إضافة ملف الإكسل الذي تريد البحث فيه')
        return
        
    worksheets = []
    for sheets in loadedWorkBook._sheets:
        worksheets.append(sheets)

    searchResults = []

    i = 0
    for sheets in loadedWorkBook._sheets:
        for sheet in sheets:
            for cell in sheet:
                try:
                    dfdf = cell.value
                    if(cell.value.find(studentName) != -1):
                        cellNumber = (RemoveStrings(cell.coordinate))
                        studentClass = loadedWorkBook.sheetnames[i]

                        name = worksheets[i]['B'+cellNumber].value
                        studentTotal = worksheets[i]['C'+cellNumber].value
                        studentPercentage = worksheets[i]['D' +
                                                          cellNumber].value
                        studentRank = worksheets[i]['E'+cellNumber].value

                        searchResults.append(Student(name, studentClass, studentTotal,
                                                     studentRank, studentPercentage))

                except (AttributeError, TypeError, ValueError):
                    continue
        i += 1

    DisplaySearchResult(searchResults)


def DisplaySearchResult(searchResults):
    ClearTab(searchStudentsTab)
    DisplaySerachScreen()
    global rowsNumber
    rowsNumber += 2

    tree = ttk.Treeview(searchStudentsTab)
    tree["columns"] = ("1", "2", "3", "4", "5")
    tree.column("1", minwidth=0, width=150, stretch=NO)
    tree.column("2", minwidth=0, width=50, stretch=NO)
    tree.column("3", minwidth=0, width=50, stretch=NO)
    tree.column("4", minwidth=0, width=75, stretch=NO)
    tree.column("5", minwidth=0, width=50, stretch=NO)

    tree.heading("1", text="اسم الطالب")
    tree.heading("2", text="الصف")
    tree.heading("3", text="المجموع")
    tree.heading("4", text="النسبة المئوية")
    tree.heading("5", text="الترتيب")

    i = 0
    for student in searchResults:
        tree.insert('', i,
                    values=(student.name, student.classN,
                            student.total, student.percentage, student.rank))
        i = i + 1

    vsb = Scrollbar(searchStudentsTab, orient="vertical", command=tree.yview)
    vsb.place(relx=0.078, rely=0.625, relheight=0.373, relwidth=0.020)
    tree.configure(yscrollcommand=vsb.set)

    tree.grid(row=rowsNumber, column=0, columnspan=100)


# DisplaySerachScreen()
# notebook.select(tab4_SearchStudents)
# ROOT.mainloop()
