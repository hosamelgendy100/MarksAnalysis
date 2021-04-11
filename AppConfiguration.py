from tkinter import *
from tkinter import ttk
import DataBase
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


class Colors:
    def __init__(self):
        self.txtColor = '#00373F'
        self.bg = '#f2f2f2'
        self.bg2 = '#9bc472'
        self.danger = '#f5beb4'


colors = Colors()
originalWorkbook = None

def SetOriginalWorkbook(workbook):
    global originalWorkbook
    originalWorkbook = workbook

def GetOriginalWorkbook():
    return originalWorkbook


def ClearTab(data):
    for widget in data.winfo_children():
        widget.destroy()


xPadding = 25
yPadding = 10
FONT_SETTINGS = ("Times New Roman", 18, 'bold')
DEFAULT_VALUES = DataBase.GetDataFromTable('DefaultValues')
CELL_FONT_SETTINGS = Font(size=12, bold=True)
CELL_FILL_SETTINGS = PatternFill(patternType="solid", start_color="DAEEF3")
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')


def GET_SUBJECTS_FROM_DB():
    return DataBase.GetDataFromTable('Subjects')


# Tkinter Configuration
ROOT = Tk()
ROOT.title("تحليل درجات الطلاب")
ROOT.geometry("900x900")
style = ttk.Style(ROOT)

style.theme_create("MyStyle", settings={
    "TNotebook": {"configure": {"tabmargins": [1, 2, 1, 1], "background": colors.bg}},
    "TNotebook.Tab": {"configure": {"padding": [5, 10], "background": colors.bg2,
                                    "foreground": 'white', "font": ('Rockwell', '16', 'bold')}}
})

style.theme_use('MyStyle')

notebook = ttk.Notebook(ROOT)
notebook.pressed_index = None


tab1_Welcome = Frame(notebook, background=colors.bg)
tab2_AddInformation = Frame(notebook, background=colors.bg)
tab3_Subjects = Frame(notebook, background=colors.bg)
tab4_SearchStudents = Frame(notebook, background=colors.bg)

tab1_Welcome.pack(fill=BOTH, expand=1)
tab2_AddInformation.pack(fill=BOTH, expand=1)
tab3_Subjects.pack(fill=BOTH, expand=1)
tab4_SearchStudents.pack(fill=BOTH, expand=1)
notebook.pack(fill=BOTH, expand=1)


notebook.add(tab1_Welcome, text=f'{"تعليمات":^35}')
notebook.add(tab2_AddInformation, text=f'{" معلومات":^35}')
notebook.add(tab3_Subjects, text=f'{"المواد":^35}')
notebook.add(tab4_SearchStudents, text=f'{"البحث ":^35}')


# Add Scroll TO Tab1
welcomCanvas = Canvas(
    tab1_Welcome, relief='raised', background=colors.bg)
welcomCanvas.pack(side=LEFT, fill=BOTH, expand=1)

myScroll_0 = Scrollbar(tab1_Welcome, orient=VERTICAL,
                       command=welcomCanvas.yview)
myScroll_0.pack(side=RIGHT, fill=Y)

welcomCanvas.configure(yscrollcommand=myScroll_0.set)
welcomCanvas.bind('<Configure>', lambda e: welcomCanvas.configure(
    scrollregion=welcomCanvas.bbox("all")))

welcomeTab = Frame(welcomCanvas, background=colors.bg)
welcomCanvas.create_window(0, 0, window=welcomeTab, anchor="nw")

###


# Add Scroll TO Tab2
addSubjectsCanvas = Canvas(
    tab3_Subjects, relief='raised', background=colors.bg)
addSubjectsCanvas.pack(side=LEFT, fill=BOTH, expand=1)

myScroll_1 = Scrollbar(tab3_Subjects, orient=VERTICAL,
                       command=addSubjectsCanvas.yview)
myScroll_1.pack(side=RIGHT, fill=Y)

addSubjectsCanvas.configure(yscrollcommand=myScroll_1.set)
addSubjectsCanvas.bind('<Configure>', lambda e: addSubjectsCanvas.configure(
    scrollregion=addSubjectsCanvas.bbox("all")))

AddSubjectsTab = Frame(addSubjectsCanvas, background=colors.bg)
addSubjectsCanvas.create_window(0, 0, window=AddSubjectsTab, anchor="nw")

###


# Add Scroll TO Tab3
searchStudentsCanvas = Canvas(tab4_SearchStudents, relief='raised')
searchStudentsCanvas.pack(side=LEFT, fill=BOTH, expand=1)

myScrol_2 = Scrollbar(tab4_SearchStudents, orient=VERTICAL,
                      command=searchStudentsCanvas.yview)
myScrol_2.pack(side=RIGHT, fill=Y)

searchStudentsCanvas.configure(yscrollcommand=myScrol_2.set)
searchStudentsCanvas.bind('<Configure>', lambda e: searchStudentsCanvas.configure(
    scrollregion=searchStudentsCanvas.bbox("all")))

searchStudentsTab = Frame(searchStudentsCanvas, background=colors.bg)
searchStudentsCanvas.create_window(0, 0, window=searchStudentsTab, anchor="nw")

###


# # Full Screen Code
# ROOT.attributes("-fullscreen", True)
# ROOT.bind("<F11>", lambda event: ROOT.attributes("-fullscreen",
#                                                  not ROOT.attributes("-fullscreen")))

# ROOT.bind("<Escape>", lambda event: ROOT.attributes("-fullscreen", False))
