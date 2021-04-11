import os
from openpyxl import  load_workbook
from tkinter import *
from tkinter import messagebox, filedialog
from AppConfiguration import  tab3_Subjects, ClearTab, DEFAULT_VALUES, SetOriginalWorkbook, tab2_AddInformation, colors, notebook, colors

fontSettings = ("Times New Roman", 16, 'bold')
txtWidth = 10

titlesLabels = []
titlesTxtBoxs = []
titlesRow = 4
subjectRow = 0
originalWorkbook = None
infoColumn = 5


def DisplayInformationScreen():
    global titlesRow
    # Delete All widgets before adding new ones
    ClearTab(tab2_AddInformation)

    # label Title
    label1 = Label(tab2_AddInformation, text="-- إدخال معلومات المدرسة --", background=colors.bg, foreground=colors.bg2,
                   padx=80, pady=15, relief=RIDGE,  font=("Times New Roman", 30, 'bold'))
    label1.grid(row=0, column=1, pady=15, columnspan=10, padx=250)

    btnAddExcelFile = Button(tab2_AddInformation, text=" +  إضافة ملف الإكسل المراد تحليله ",
                             font=fontSettings, bg=colors.bg2, foreground="white", border=2,
                              cursor="hand2",command=lambda: AddExcelFile())
    btnAddExcelFile.grid(row=1, column=infoColumn, pady=5,columnspan=3)

    # label Title
    label1 = Label(tab2_AddInformation, text=" هذه القيم افتراضية كمثال فقط لإدخال البيانات، عدل هذه القيم بناءً على الملف الخاص بك",
                   background=colors.bg, pady=15, font=fontSettings, wraplength=400)
    label1.grid(row=3, column=infoColumn, pady=5, columnspan=3)

    titlesLabels.append(":اسم الوزارة")
    titlesLabels.append(":اسم المدرسة")
    titlesLabels.append(":عنوان الملف")
    titlesLabels.append("أول خلية لأسماء الطلاب")
    titlesLabels.append("أول خلية لصفوف الطلاب")
    titlesLabels.append("عدد المراكز الأولى")
    titlesLabels.append("الدرجة النهائية للمادة")

    columnNumber = 0
    # add Document Info txtboxs
    for i in range(3):
        titleTxtBox = Entry(tab2_AddInformation,
                            font=fontSettings, width=txtWidth*6)
        titleTxtBox.configure(justify='right')
        titleTxtBox.grid(row=titlesRow, column=columnNumber,pady=10, columnspan=10)
        titleTxtBox.insert(0, DEFAULT_VALUES[i][2])
        titlesTxtBoxs.append(titleTxtBox)

        label = Label(tab2_AddInformation, text=titlesLabels[i], background=colors.bg,
                      foreground=colors.txtColor, font=fontSettings)
        label.grid(row=titlesRow, column=columnNumber + 8, sticky='E')

        titlesRow += 2

    columnNumber += 5
    # add Students info text boxs
    for i in range(3, len(titlesLabels)):
        if(i % 2 != 0):
            titlesRow += 1
            columnNumber = 5

        titleTxtBox = Entry(tab2_AddInformation,font=fontSettings, width=txtWidth)
        titleTxtBox.grid(row=titlesRow, column=columnNumber, pady=10)
        titleTxtBox.insert(0, DEFAULT_VALUES[i][2])
        titlesTxtBoxs.append(titleTxtBox)

        label = Label(tab2_AddInformation, text=titlesLabels[i], background=colors.bg,
                      foreground=colors.txtColor, font=fontSettings)
        label.grid(row=titlesRow, column=columnNumber + 1, sticky='E')

        columnNumber += 2

    titlesRow += 1
    btnAddNewSubject = Button(
        tab2_AddInformation, text=" بيانات المواد الدراسية -> ", font=fontSettings, bg=colors.bg2,
        foreground="white", border=2, cursor="hand2",
        command=lambda: BtnShowSubjectWindow())
    btnAddNewSubject.grid(
        row=titlesRow, column=infoColumn, pady=1, columnspan=3)


def BtnShowSubjectWindow():
    #originalWorkbook = GetOriginalWorkbook()
    if(originalWorkbook == None):
        messagebox.showerror('Error', 'برجاء تحميل ملف الإكسل المراد تحليله')
        return

    notebook.tab(2, state="normal")
    notebook.select(tab3_Subjects)


excelFileTitle = None
def AddExcelFile():
    global originalWorkbook
    global excelFileTitle
    if(excelFileTitle != None):excelFileTitle.destroy()

    importFilePath = filedialog.askopenfilename(
        filetypes=[("Excel files", "*.xlsx")])
    onlyFileName = os.path.split(importFilePath)
    excelFileTitle = Label(tab2_AddInformation, text='File: ' + onlyFileName[1], background=colors.bg,
                  foreground=colors.txtColor, font=fontSettings)
    excelFileTitle.grid(row=2, column=infoColumn, pady=5, columnspan=3)

    originalWorkbook = load_workbook(importFilePath)
    SetOriginalWorkbook(originalWorkbook)

