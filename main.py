from Student import Student
from openpyxl import Workbook, load_workbook
import arabic_reshaper
from bidi.algorithm import get_display
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def ArabicPrint(valueToPrint):
    text = arabic_reshaper.reshape(valueToPrint)
    print(get_display(text))


def RemoveStrings(sequance):
    seq_type = type(sequance)
    return seq_type().join(filter(seq_type.isdigit, sequance))


def RemoveNumbers(word):
    regex = re.compile('[^a-zA-Z]')
    return regex.sub('', word)


cellFontSettings = Font(size=12, bold=True)
cellFillSettings = PatternFill(patternType="solid", start_color="DAEEF3")


# Import Workbook
workbook = load_workbook('8th.xlsx')
orginalSheet = workbook.active

nameCN = "B5"
classCN = "C5"
schoolName = "وزارة التعليم والتعليم العالي - مدرسة معيذر الابتدائية للبنين"
sheetTitle = "تحليل نتائج اختبارات منتصف الفصل الثاني 2020-2021"
bestStudentsMax = 20
bestStudentsSheet = False

classCellValue = orginalSheet[classCN].value

# get the first cell number
cellNumber = int(RemoveStrings(nameCN))-1
subjectsCN = ["E5", "G5", "I5", "K5", "M5", "O5", "Q5"]
subjectsMarks = [15, 15, 15, 15, 15, 15, 15]

# creats new Excel File
finishedWorkbook = Workbook()
# Removes the default created workbook
finishedWorkbook.remove(finishedWorkbook.active)


def getSubjectsTotal():
    total: float = 0.0
    for subject in subjectsMarks:
        total += subject
    return total


subjectsTotal = getSubjectsTotal()


def CalcStudentTotal(subjects):
    studentName = orginalSheet[nameCN].value
    studentClass = orginalSheet[classCN].value

    total: float = 0.0
    for i in range(len(subjects)):
        subjects[i] = RemoveNumbers(subjects[i])
        subjects[i] = subjects[i] + str(cellNumber)
        total += orginalSheet[subjects[i]].value

    GetBestStudents(studentName, studentClass, total)
    return total


bestStudentsList = []
def GetBestStudents(studentName, studentClass, total):
    global bestStudentsList
    if(len(bestStudentsList) < bestStudentsMax):
        bestStudentsList.append(Student(studentName, studentClass, total))
    else:
        bestStudentsList = sorted(
            bestStudentsList, key=lambda x: x.total, reverse=False)

        if(total > bestStudentsList[0].total):
            bestStudentsList.remove(bestStudentsList[0])
            bestStudentsList.append(Student(studentName, studentClass, total))

        elif(total == bestStudentsList[0].total):
            bestStudentsList.append(Student(studentName, studentClass, total))


def AddNewSheet(name):
    global finishedWorkbook
    sheetName = str(name).replace('/', '-')
    newSheet = finishedWorkbook.create_sheet(sheetName)
    newSheet.sheet_view.rightToLeft = True
    CreateColumnNames(newSheet)

    return newSheet


def CreateColumnNames(newSheet):
    # add School Name
    newSheet.merge_cells('A1:E1')
    newSheet.cell(1, 1, schoolName)

    # add  sheet title
    newSheet.merge_cells('A2:E2')
    newSheet.cell(2, 1, sheetTitle)

    # add  class title
    if (bestStudentsSheet == False):
        newSheet.merge_cells('A3:E3')
        newSheet.cell(3, 1, 'الصف   ' + str(orginalSheet[classCN].value))
    else:
        newSheet.merge_cells('A3:E3')
        newSheet.cell(3, 1, 'الأوائل على جميع الصفوف')

    alignSettings = Alignment(horizontal='center', vertical='center')

    newSheet.cell(1, 1).alignment = alignSettings
    newSheet.cell(2, 1).alignment = alignSettings
    newSheet.cell(3, 1).alignment = alignSettings

    newSheet.row_dimensions[1].height = 40

    newSheet.cell(1, 1).font = Font(size=14, bold=True, color="0000FF")
    newSheet.cell(2, 1).font = Font(size=14, bold=True, color="FF0000")
    newSheet.cell(3, 1).font = Font(size=14, bold=True, color="0000FF")

    newSheet["A4"] = "م"
    newSheet["A4"].font = cellFontSettings
    newSheet["A4"].fill = cellFillSettings

    newSheet["B4"] = "الإسم"
    newSheet["B4"].font = cellFontSettings
    newSheet["B4"].fill = cellFillSettings

    newSheet["C4"] = "المجموع"
    newSheet["C4"].font = cellFontSettings
    newSheet["C4"].fill = cellFillSettings

    newSheet["D4"] = "النسبة %"
    newSheet["D4"].font = cellFontSettings
    newSheet["D4"].fill = cellFillSettings

    newSheet["E4"] = "الترتيب"
    newSheet["E4"].font = cellFontSettings
    newSheet["E4"].fill = cellFillSettings


def AdjustSheetWidth(newSheet):
    newSheet.column_dimensions['A'].width = 5
    newSheet.column_dimensions['B'].width = 40
    newSheet.column_dimensions['E'].width = 7


def set_borders(ws, cell_range):
    thin = Side(border_style="thin", color="000000")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)


def AnalyzeData():
    global nameCN
    global classCN
    global cellNumber
    global classCellValue
    global newSheetCellNumber
    newSheetCellNumber = 4

    # Save Workbook
    newSheet = AddNewSheet(orginalSheet[classCN].value)
    newSheet.sheet_view.rightToLeft = True

    while(orginalSheet[nameCN].value != None):
        nameCN = RemoveNumbers(nameCN)
        classCN = RemoveNumbers(classCN)

        # Incerment Excel
        cellNumber = int(cellNumber) + 1
        nameCN = nameCN + str(cellNumber)
        classCN = classCN + str(cellNumber)

        if(orginalSheet[nameCN].value == None):
            AdjustSheetWidth(newSheet)
            set_borders(newSheet, 'A4:E'+str(newSheetCellNumber))
            AddRank(newSheet, newSheetCellNumber)
            return

        # Add New Tab
        if(classCellValue != orginalSheet[classCN].value):
            classCellValue = orginalSheet[classCN].value
            AdjustSheetWidth(newSheet)
            set_borders(newSheet, 'A4:E'+str(newSheetCellNumber))
            AddRank(newSheet, newSheetCellNumber)
            newSheet = AddNewSheet(orginalSheet[classCN].value)
            newSheetCellNumber = 4

        # Save data to excel workbook
        newSheetCellNumber += 1
        newSheet["A"+str(newSheetCellNumber)] = str(newSheetCellNumber-4)
        newSheet["A"+str(newSheetCellNumber)].font = cellFontSettings
        newSheet["A"+str(newSheetCellNumber)].fill = cellFillSettings
        newSheet["A"+str(newSheetCellNumber)
                 ].alignment = Alignment(horizontal='right')

        # Save Student Name
        newSheet["B"+str(newSheetCellNumber)] = str(orginalSheet[nameCN].value)
        newSheet["B"+str(newSheetCellNumber)].font = cellFontSettings

        # Save Student Total
        studentTotal = CalcStudentTotal(subjectsCN)
        newSheet["C"+str(newSheetCellNumber)] = float(studentTotal)
        # convert cell to number
        newSheet["C"+str(newSheetCellNumber)].number_format = '0.0'

        newSheet["C"+str(newSheetCellNumber)].font = cellFontSettings
        newSheet["C"+str(newSheetCellNumber)
                 ].alignment = Alignment(horizontal='right')

        # Save Student Percentage
        newSheet["D"+str(newSheetCellNumber)] = str(
            round(studentTotal*100/subjectsTotal, 2))+' %'
        newSheet["D"+str(newSheetCellNumber)].font = cellFontSettings
        newSheet["D"+str(newSheetCellNumber)
                 ].alignment = Alignment(horizontal='right')




def AddRank2(newSheet, lastValue):
    for i in range(5, lastValue+1):
        newSheet['E'+str(i)] = "=RANK(C{i},$C$5:$C${lastValue},0)".format(
            i=i, lastValue=lastValue)
        newSheet['E'+str(i)].font = cellFontSettings



def AddRank(newSheet, lastValue):
  totalList=[]
  for i in range(5, lastValue+1):
    totalList.append(newSheet['C'+str(i)].value)
  
  totalList=sorted(totalList,reverse=True)

  rankList=[]
  rankVal=1
  rankList.append(rankVal)

  print(orginalSheet[classCN].value)
  for i in range(0,len(totalList)):
    print(totalList[i])
  print('------------------------------------')
  #   if(totalList[i]<totalList[i-1]):
  #     rankVal +=1
  #     rankList[i] =rankVal
  #   else:
  #     rankList[i] =rankVal

  # for i in range(5, lastValue+1):
  #       newSheet['E'+str(i)] = rankList[i-5]
  #       newSheet['E'+str(i)].font = cellFontSettings

        


def DisplayBestStudents():
    global bestStudentsList
    global bestStudentsSheet
    bestStudentsList = sorted(
        bestStudentsList, key=lambda x: x.total, reverse=True)

    newSheet = finishedWorkbook.create_sheet("الأوائل")
    newSheet.sheet_view.rightToLeft = True
    bestStudentsSheet = True
    CreateColumnNames(newSheet)
    AdjustSheetWidth(newSheet)

    newSheet["F4"] = "الصف"
    newSheet["F4"].font = cellFontSettings
    newSheet["F4"].fill = cellFillSettings

    for i in range(5, len(bestStudentsList)+5):
        newSheet['A'+str(i)] = str(i-4)
        newSheet['B'+str(i)] = str(bestStudentsList[i-5].name)

        newSheet['C'+str(i)] = int(bestStudentsList[i-5].total)
        newSheet['C'+str(i)].number_format = '0'

        newSheet['D'+str(i)] = str(round(bestStudentsList[i -
                                                          5].total*100/subjectsTotal, 2))+' %'
        newSheet['F'+str(i)] = str(bestStudentsList[i-5].classN)

    AddRank(newSheet, i)
    set_borders(newSheet, 'A4:F'+str(i))


AnalyzeData()
DisplayBestStudents()
finishedWorkbook.save(filename="Analysis.xlsx")

#os.startfile('Analysis.xlsx')
