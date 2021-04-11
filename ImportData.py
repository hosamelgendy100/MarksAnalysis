from AppConfiguration import ALIGN_CENTER, CELL_FILL_SETTINGS, CELL_FONT_SETTINGS
from ExcelOperations import AddAverage, AdjustSheetWidth, RemoveNumbers, RemoveStrings, SaveExcelFile, SetAlignRange, SetBorders, SetFillRange, SetFontRange, getSubjectsTotal
import os
import webbrowser
from Classes import Student, Rank
from openpyxl import Workbook
import re
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import pathlib



ministryName = None
schoolName = None
sheetTitle = None
studentNameCN = None
studentClassCN = None
bestStudentsMax = None
finishedWorkbook = None
subjectsCN = None
subjectsFullMark = None
bestStudentsSheet = False
classCellValue = None
# get the first cell number
firstCellNumber = None

# change this number if you need to add more titles
titleRow = str(5)
newSheet = None





def GetDataToAnalyse(_originalWorkbook, _ministryName, _schoolName, _sheetTitle,
                     _studentNameCN, _studentClassCN, _bestStudentsMax, _subjectsCN, _subjectsFullMark):

    global originalWorkbook
    global ministryName
    global schoolName
    global sheetTitle
    global studentNameCN
    global studentClassCN
    global bestStudentsMax
    global classCellValue
    global firstCellNumber
    global orginalSheet
    global subjectsCN
    global subjectsFullMark

    originalWorkbook = _originalWorkbook
    ministryName = _ministryName
    schoolName = _schoolName
    sheetTitle = _sheetTitle
    studentNameCN = _studentNameCN
    studentClassCN = _studentClassCN
    bestStudentsMax = _bestStudentsMax
    subjectsCN = _subjectsCN
    subjectsFullMark = _subjectsFullMark

    AnalyzeData()



def CalcStudentTotal(subjects):
    studentName = orginalSheet[studentNameCN].value
    studentClass = orginalSheet[studentClassCN].value

    total: float = 0
    for i in range(len(subjects)):
        subjects[i] = RemoveNumbers(subjects[i])
        subjects[i] = subjects[i] + str(firstCellNumber)
        total += orginalSheet[subjects[i]].value

    GetBestStudents(studentName, studentClass, total)
    return total


bestStudentsList = []


def GetBestStudents(studentName, studentClass, total: float):
    global bestStudentsList
    if(len(bestStudentsList) < bestStudentsMax):
        bestStudentsList.append(
            Student(studentName, studentClass, total))
    else:
        bestStudentsList = sorted(
            bestStudentsList, key=lambda x: x.total, reverse=False)

        if(total > bestStudentsList[0].total):
            bestStudentsList.remove(bestStudentsList[0])
            bestStudentsList.append(
                Student(studentName, studentClass, total, 2))

        elif(total == bestStudentsList[0].total):
            bestStudentsList.append(
                Student(studentName, studentClass, total, 2))


def AddNewSheet(name, finishedWorkbook):
    sheetName = str(name).replace('/', '-')
    newSheet = finishedWorkbook.create_sheet(sheetName)
    newSheet.sheet_view.rightToLeft = True
    CreateColumnNames(newSheet)

    return newSheet


def CreateColumnNames(newSheet):
    titleRow = str(5)

    # add Ministry Name
    newSheet.merge_cells('A1:E1')
    newSheet.cell(1, 1, ministryName)

    # add School Name
    newSheet.merge_cells('A2:E2')
    newSheet.cell(2, 1, schoolName)

    # add  sheet title
    newSheet.merge_cells('A3:E3')
    newSheet.cell(3, 1, sheetTitle)

    # add  class title
    newSheet.merge_cells('A4:E4')
    if (bestStudentsSheet == False):
        newSheet.cell(4, 1, 'الصف   ' +
                      str(orginalSheet[studentClassCN].value))
    else:
        newSheet.cell(4, 1, 'الأوائل على جميع الصفوف')

    SetAlignRange(newSheet,"A1:A4", ALIGN_CENTER)
    newSheet.row_dimensions[1].height = 40

    newSheet.cell(1, 1).font = Font(size=14, bold=True, color="0000FF")
    newSheet.cell(2, 1).font = Font(size=14, bold=True, color="0000FF")
    newSheet.cell(3, 1).font = Font(size=14, bold=True, color="FF0000")
    newSheet.cell(4, 1).font = Font(size=14, bold=True, color="FF0000")

    newSheet["A"+titleRow] = "م"
    newSheet["B"+titleRow] = "الإسم"
    newSheet["C"+titleRow] = "المجموع"
    newSheet["D"+titleRow] = "النسبة %"
    newSheet["E"+titleRow] = "الترتيب"
   
    SetFillRange(newSheet,"A"+titleRow + ":E"+titleRow)
    SetFontRange(newSheet, "A"+titleRow + ":E"+titleRow)

    titleRow = int(titleRow)






def AnalyzeData():
    bestStudentsList.clear()
    global studentNameCN
    global studentClassCN
    global firstCellNumber
    global classCellValue
    global newSheetCellNumber
    global finishedWorkbook
    global orginalSheet
    global subjectsTotal
    global originalWorkbook
    global newSheet

    newSheetCellNumber = int(titleRow)

    # Import Workbook
    orginalSheet = originalWorkbook.active
    classCellValue = orginalSheet[studentClassCN].value
    firstCellNumber = int(RemoveStrings(studentNameCN))-1

    # creats new Excel File
    finishedWorkbook = Workbook()
    # Removes the default created workbook
    finishedWorkbook.remove(finishedWorkbook.active)

    subjectsTotal = getSubjectsTotal(subjectsFullMark, subjectsCN)

    # add New sheet by the class name
    newSheet = AddNewSheet(
        orginalSheet[studentClassCN].value, finishedWorkbook)
    
    
    while(orginalSheet[studentNameCN].value != None):
        studentNameCN = RemoveNumbers(studentNameCN)
        studentClassCN = RemoveNumbers(studentClassCN)

        # Incerment Excel
        firstCellNumber = int(firstCellNumber) + 1
        studentNameCN = studentNameCN + str(firstCellNumber)
        studentClassCN = studentClassCN + str(firstCellNumber)

        if(orginalSheet[studentNameCN].value == None):
            AdjustSheetWidth(newSheet)
            SetBorders(newSheet, 'A'+str(titleRow) +
                        ':E'+str(newSheetCellNumber+2))
            AddRank(newSheet, newSheetCellNumber)
            AddAverage(newSheet, newSheetCellNumber, titleRow)
            continue

        # Add New Tab
        if(classCellValue != orginalSheet[studentClassCN].value):
            classCellValue = orginalSheet[studentClassCN].value
            AdjustSheetWidth(newSheet)
            SetBorders(newSheet, 'A'+str(titleRow) +
                        ':E'+str(newSheetCellNumber+2))

            AddAverage(newSheet, newSheetCellNumber, titleRow)

            AddRank(newSheet, newSheetCellNumber)
            newSheet = AddNewSheet(
                orginalSheet[studentClassCN].value, finishedWorkbook)
            newSheetCellNumber = int(titleRow)
            print('creating sheet for class' +
                  str(orginalSheet[studentClassCN].value))

        # Save data to excel workbook
        newSheetCellNumber = str(int(newSheetCellNumber) + 1)
        newSheet["A"+newSheetCellNumber] = str(int(newSheetCellNumber)-int(titleRow))
        newSheet["A"+newSheetCellNumber].fill = CELL_FILL_SETTINGS

        # Save Student Name
        newSheet["B"+newSheetCellNumber] = str(orginalSheet[studentNameCN].value)

        # Save Student Total
        studentTotal = CalcStudentTotal(subjectsCN)
        newSheet["C"+newSheetCellNumber] = round(studentTotal, 2)
        # convert cell to number
        newSheet["C"+newSheetCellNumber].number_format = '0.00'

        newSheet["C"+newSheetCellNumber
                 ].alignment = ALIGN_CENTER

        # Save Student Percentage
        newSheet["D"+newSheetCellNumber] = (str(round(studentTotal*100/subjectsTotal, 2))+'%')
        newSheet["D"+newSheetCellNumber
                 ].style = "Percent"

        newSheet["D"+str(newSheetCellNumber)].alignment = ALIGN_CENTER
        SetFontRange(newSheet, "A"+ newSheetCellNumber +":D"+newSheetCellNumber)
        newSheetCellNumber = int(newSheetCellNumber)

    AddBestStudents()
    SaveExcelFile(finishedWorkbook, sheetTitle)





def AddRank(newSheet, lastValue):
    totalList = []
    for i in range(int(titleRow)+1, lastValue+1):
        totalList.append(newSheet['C'+str(i)].value)

    totalList = sorted(totalList, reverse=True)

    rankList = []
    rankNum = 1
    rankList.append(Rank(totalList[0], rankNum))

    for i in range(1, len(totalList)):
        if(totalList[i] < totalList[i-1]):
            rankNum += 1

        rankList.append(Rank(totalList[i], rankNum))

    for i in range(int(titleRow)+1, int(lastValue)+1):
        rankMatch = next((x for x in rankList if x.key ==
                         newSheet['C'+str(i)].value), None)
        newSheet['E'+str(i)] = rankMatch.value
        newSheet['E'+str(i)].font = CELL_FONT_SETTINGS


def AddBestStudents():
    global bestStudentsList
    global bestStudentsSheet
    global titleRow

    bestStudentsList = sorted(
        bestStudentsList, key=lambda x: x.total, reverse=True)

    newSheet = finishedWorkbook.create_sheet("الأوائل")
    newSheet.sheet_view.rightToLeft = True
    bestStudentsSheet = True
    CreateColumnNames(newSheet)
    AdjustSheetWidth(newSheet)

    newSheet["F"+str(titleRow)] = "الصف"
    newSheet["F"+str(titleRow)].font = CELL_FONT_SETTINGS
    newSheet["F"+str(titleRow)].fill = CELL_FILL_SETTINGS

    titleRow =int(titleRow)
    for i in range(titleRow+1, len(bestStudentsList)+titleRow):
        i=str(i)
        newSheet['A'+i] = str(int(i)-5)
        newSheet['A'+i].fill = CELL_FILL_SETTINGS
        newSheet['A'+i].alignment = Alignment(horizontal='right')
        newSheet['B'+i] = str(bestStudentsList[int(i)-titleRow].name)
        newSheet['C'+i] = float(bestStudentsList[int(i)-titleRow].total)
        newSheet['C'+i].number_format = '0.00'
        newSheet['D'+i] = str(round(bestStudentsList[int(i) -
                                                   titleRow].total*100/subjectsTotal, 2))+' %'
        newSheet['F'+i] = str(bestStudentsList[int(i)-titleRow].classN)

        SetFontRange(newSheet, "A"+i + ":F"+i)
        i=int(i)

    AddRank(newSheet, i)
    SetBorders(newSheet, 'A'+str(titleRow)+':F' + str(i))
